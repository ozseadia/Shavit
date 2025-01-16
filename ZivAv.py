# -*- coding: utf-8 -*-
"""
Created on Tue Nov 17 10:03:20 2020

@author: OzSea
"""
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from collections import Counter
from datetime import datetime
from datetime import date
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import filedialog
import time
import streamlit as st
#from streamlit import caching
import os
import numpy as np
from io import BytesIO
redfill = PatternFill(fill_type='solid', start_color='00FF0000', end_color='00FF0000')
greenfill = PatternFill(fill_type='solid', start_color='0000FF00', end_color='0000FF00')
grayfill = PatternFill(fill_type='solid', start_color='00C0C0C0', end_color='00C0C0C0')
yellowfill = PatternFill(fill_type='solid', start_color='00FFFF00', end_color='00FFFF00')
violetfill = PatternFill(fill_type='solid', start_color='00CCCCFF', end_color='00CCCCFF')
#FileName=r"G:\Oz\Shavit\Master - orders status\example files\Open PO_s.csv"
#FileName=r"G:\Oz\Shavit\Master - orders status\example files\הזמנות לפי לקוח-82.xlsx"
@st.cache_data()
def LoadData(FileName):
    #load the data
    #'''
    #path='G:\\Oz\\Bursa\\Insider\\'
    # 'merged_notadj170820.xlsx' #'merged_adj1908.xlsx' #'merged060820.xlsx'
    #FileName='merged_V6A_0709.xlsx' #'merged_adj1908.xlsx'
    try :
        data=pd.read_csv(FileName,skiprows=(0)) #for landa csv file
    except: 
        1 
    try :
        data=pd.read_excel(FileName)
    except:
        0
    return (data)
@st.cache_data()
def SelectFile():
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    file_path = filedialog.askopenfilename()
    #temp=file_path.split('/')
    #temp[-1]=SN+'.xlsx'
    #Result_file_path='/'.join(temp)
    root.destroy()
    return (file_path)
#@st.cache_data()
def main (QT,C,Customer,file_path_C):

    L=C
    
    # QT.keys()[4] :  הזמנת רכש לקוח 
    # L.keys()[0]  : Order  : לקוח לנדא
    
    '''
    this function get:
    QT :Qtouch report
    Co:stumer report
    Master Report
    '''
    
    today = date.today()
    today=datetime.strptime(str(today),'%Y-%m-%d')
    #-------------------------------------------------------------------------------------------
    #collecting customer data  Qty >0
    #-------------------------------------------------------------------------------------------
    MasterFileName=r'G:\Oz\Shavit\Master - orders status\example files\Status Report Master.xlsx'
    #wb =openpyxl.load_workbook(file_path_C)
    wb=Workbook()
    ws=wb.active
    for r in dataframe_to_rows(C, index=True, header=True):
        ws.append(r)
    head_tail = os.path.split(MasterFileName)
    SavedFilePath=head_tail[0]+'\\' + Customer + str(date.today()) + '.xlsx'
    #order number + line number + PartNumber
    List=list((QT[QT.keys()[3]].astype(str))+(QT[QT.keys()[4]].astype(str)))
    print(List)
    #ws.cell(1,11).value='תאריך אספקה'
    cl=5
    r=4
    ws.cell(1+r,12+cl).value='הערות'
    ws.cell(1+r,13+cl).value='יתרה לאספקה'
    ws.cell(1+r,14+cl).value='מספר הזמנה שביט'
    ws.cell(1+r,15+cl).value='מק"ט'
    ws.cell(1+r,16+cl).value='הפרש בימים'
    r=3
    for i in range(len(L)):
        #ind=[]
        #element=str(L[L.keys()[2]][i])+str(L[L.keys()[3]][i])
        element=str(L[L.keys()[1]][i])+str(L[L.keys()[4]][i])
        print(element)
        
        try:
            ind=[]
            try:
                ind=List.index(element)
            except:
                0
            if ind!=[]:
                print(ind)
                ws.cell(i+r,11+cl).value=QT[QT.keys()[10]][ind] # תאריך אספקה
                ws.cell(i+r,11+cl).number_format = 'DD/MM/YYYY'
                temp=np.busday_count(str(today.date()),str((QT[QT.keys()[10]][ind]).date()),weekmask=[1,1,1,1,1,0,1])
                temp1=((QT[QT.keys()[10]][ind]).date().year==today.date().year)*((QT[QT.keys()[10]][ind]).date().month==today.date().month)
                print(temp1)
                ws.cell(i+r,16+cl).value=temp
                if temp>0:
                    if temp1:
                        for j in range(1,17+cl):ws.cell(i+r,j).fill=violetfill
                    else: 
                        for j in range(1,17+cl):ws.cell(i+r,j).fill=greenfill
                else:
                    for j in range(1,17+cl):ws.cell(i+r,j).fill=redfill
                ws.cell(i+r,12+cl).value=QT[QT.keys()[11]][ind] # סטאטוס
                ws.cell(i+r,13+cl).value=QT[QT.keys()[9]][ind] # יתרה לאספקה
                ws.cell(i+r,14+cl).value=QT[QT.keys()[2]][ind] #מספר הזמנה שביט
                ws.cell(i+r,15+cl).value=QT[QT.keys()[4]][ind] #מקט
                    #ws.cell(i,20).value=QT[QT.keys()[11]][ind] # תאריך אספקה שביט
            else:
                if i>3:
                    for j in range(1,17+cl):ws.cell(i+r,j).fill=yellowfill
        except:
            0
    #ws.cell(1,1).value=Customer
    ws.delete_rows(1,2)
    output = BytesIO()
    #b.save(SavedFilePath)
    wb.save(output)
    # Rewind the buffer
    output.seek(0)
    st.download_button(label="Download Excel file", 
                       data=output,
                       file_name="sample.xlsx", 
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    
    '''    
    for i in range(len(L)):
        ws.cell(i+2,21).value=np.busday_count(str(today.date()),str((QT[QT.keys()[5]][indQN[i]]).date()),weekmask=[1,1,1,1,1,0,1])
        wb.save(SavedFilePath)
        if (ws.cell(indLN[i]+3,12).value):
            if ws.cell(indLN[i]+3,12).value <0:
                for j in range(1,16):ws.cell(indLN[i]+3,j).fill=redfill
            else:
                for j in range(1,16):ws.cell(indLN[i]+3,j).fill=greenfill
    '''
    #*******************************************************************************************
    '''
    #-------------------------------------------------------------------------------------------
    #collecting QTouch data QTY 
    #-------------------------------------------------------------------------------------------
    N=dict(Counter(L[L.keys()[2]]+(L[L.keys()[4]])))
    A=list(N.keys())
    for k in A:
        indLP=[i for i, e in enumerate(L[L.keys()[2]]+(L[L.keys()[4]])) if e == k and L[L.keys()[6]][i]>=0]
        indLN=[i for i, e in enumerate(L[L.keys()[2]]+(L[L.keys()[4]])) if e == k and L[L.keys()[6]][i]<0]
        indQP=[i for i, e in enumerate(QT[QT.keys()[4]]+QT[QT.keys()[6]]) if e == k and QT[QT.keys()[9]][i]>=0]
        indQN=[i for i, e in enumerate(QT[QT.keys()[4]]+QT[QT.keys()[6]]) if e == k and QT[QT.keys()[9]][i]<0]
        #-----------------------------------------------------------------------------------------------            
        #collecting QTouch data QTY >=0
        #-----------------------------------------------------------------------------------------------    
        
        for i in range(len(indLP)):
            if i<len(indQP):
                if QT[QT.keys()[9]][indQP[i]]>=0:
                    ws.cell(indLP[i]+3,9).value=QT[QT.keys()[9]][indQP[i]] # יתרה לאספקה
                    ws.cell(indLP[i]+3,11).value=QT[QT.keys()[3]][indQP[i]] # הזמנת SO פנימית
                    ws.cell(indLP[i]+3,13).value=QT[QT.keys()[5]][indQP[i]] # תאריך אספקה QT
                    #ws.cell(indLP[i]+3,12).value=(QT[QT.keys()[5]][indQP[i]]-today).days
                    ws.cell(indLP[i]+3,12).value=np.busday_count(str(today.date()),str((QT[QT.keys()[5]][indQP[i]]).date()),weekmask=[1,1,1,1,1,0,1])
                    if ws.cell(indLP[i]+3,12).value <0:
                        for j in range(1,16):ws.cell(indLP[i]+3,j).fill=redfill
                    else:
                        for j in range(1,16):ws.cell(indLP[i]+3,j).fill=greenfill
            else:
                ws.cell(indLP[i]+3,9).value='Nan' # יתרה לאספקה
                ws.cell(indLP[i]+3,11).value='Nan' # הזמנת SO פנימית
                ws.cell(indLP[i]+3,13).value='Nan' # תאריך אספקה QT
        wb.save(SavedFilePath)       
        #-----------------------------------------------------------------------------------------------            
        #collecting QTouch data QTY <0
        #-----------------------------------------------------------------------------------------------    
        for i in range(len(indLN)):
            if i<len(indQN):
                if QT[QT.keys()[9]][indQN[i]]<0:
                    ws.cell(indLN[i]+3,9).value=QT[QT.keys()[9]][indQN[i]] # יתרה לאספקה
                    ws.cell(indLN[i]+3,11).value=QT[QT.keys()[3]][indQN[i]] # הזמנת SO פנימית
                    ws.cell(indLN[i]+3,13).value=QT[QT.keys()[5]][indQN[i]] # תאריך אספקה QT
                    #ws.cell(indLN[i]+3,12).value=(QT[QT.keys()[5]][indQN[i]]-today).days
                    ws.cell(indLP[i]+3,12).value=np.busday_count(str(today.date()),str((QT[QT.keys()[5]][indQN[i]]).date()),weekmask=[1,1,1,1,1,0,1])
                    wb.save(SavedFilePath)
                    if (ws.cell(indLN[i]+3,12).value):
                        if ws.cell(indLN[i]+3,12).value <0:
                            for j in range(1,16):ws.cell(indLN[i]+3,j).fill=redfill
                        else:
                            for j in range(1,16):ws.cell(indLN[i]+3,j).fill=greenfill
            else:
                ws.cell(indLN[i]+3,9).value='Nan' # יתרה לאספקה
                ws.cell(indLN[i]+3,11).value='Nan' # הזמנת SO פנימית
                ws.cell(indLN[i]+3,13).value='Nan' # תאריך אספקה QT
                    
    wb.save(SavedFilePath)
    #-----------------------------------------------------------------------------------------------
    # MARKS
    #----------------------------------------------------------------------------------------------
    POL=dict(Counter(L[L.keys()[2]]))
    
    A=list(POL.keys())
    for k in A:
        indL=[i for i, e in enumerate(L[L.keys()[2]]) if e == k]
        indQ=[i for i, e in enumerate(QT[QT.keys()[4]]) if e == k]
        if not indQ:
            for i in range(len(indL)):
                ws.cell(indL[i]+3,15).value='PO Does Not Exist'
                for j in range(1,16):ws.cell(indL[i]+3,j).fill=yellowfill
        else:  
            for i in range(len(indL)):
                if L[L.keys()[4]][indL[i]] not in list(QT[QT.keys()[6]][indQ]) :
                    ws.cell(indL[i]+3,15).value='PN Does Not Exist In PO'
                    for j in range(1,16):ws.cell(indL[i]+3,j).fill=grayfill
                    
    wb.save(SavedFilePath)
'''     



        
   
                